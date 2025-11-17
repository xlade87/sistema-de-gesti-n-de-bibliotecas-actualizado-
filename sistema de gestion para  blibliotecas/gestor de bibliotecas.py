import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import mysql.connector
from datetime import datetime
import re
from tkcalendar import DateEntry
from PIL import Image, ImageTk
import os
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_DISPONIBLE = True
except ImportError:
    OPENPYXL_DISPONIBLE = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    REPORTLAB_DISPONIBLE = True
except ImportError:
    REPORTLAB_DISPONIBLE = False


# CONEXIÓN A BASE DE DATOS
class DatabaseConnection:
    def __init__(self):
        self.connection = None
        self.cursor = None

    def connect(self):
        try:
            self.connection = mysql.connector.connect(
                host="localhost",
                database="biblioteca_personal",
                user="root",
                password="",
                autocommit=True
            )
            self.cursor = self.connection.cursor(buffered=True)
            return True
        except mysql.connector.Error as error:
            messagebox.showerror("Error de Conexión", f"Error conectando a la base de datos: {error}")
            return False

    def disconnect(self):
        if self.cursor:
            self.cursor.close()
        if self.connection:
            self.connection.close()

    def execute_query(self, query, parameters=None):
        try:
            if parameters:
                self.cursor.execute(query, parameters)
            else:
                self.cursor.execute(query)

            if query.strip().upper().startswith('SELECT'):
                return True, self.cursor.fetchall()
            else:
                self.connection.commit()

                if query.strip().upper().startswith('CALL'):
                    result = self.cursor.fetchall()
                    self.cursor.nextset()
                    return True, result
                else:
                    return True, "Operación exitosa"
        except mysql.connector.Error as error:
            return False, str(error)


db = DatabaseConnection()


# VALIDACIONES
class Validaciones:
    @staticmethod
    def solo_numeros(texto):
        """Solo permite números"""
        return texto.isdigit() or texto == ""

    @staticmethod
    def validar_email(email):
        """Valida formato de email"""
        return re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email) is not None

    @staticmethod
    def validar_imagen(ruta):
        """Valida imagen"""
        if not ruta:
            return False, "No hay imagen"

        extensiones = ['.jpg', '.jpeg', '.png', '.gif']
        ext = os.path.splitext(ruta)[1].lower()

        if ext not in extensiones:
            return False, f"Solo: {', '.join(extensiones)}"

        # Verificar tamaño (máximo 2MB)
        tamano = os.path.getsize(ruta) / (1024 * 1024)
        if tamano > 2:
            return False, "Máximo 2MB"

        return True, "OK"


# MANEJO DE IMÁGENES
class ImagenManager:
    def __init__(self):
        self.ruta_actual = None

    def cargar_imagen_para_mostrar(self, ruta, tamaño=(150, 150)):
        """Carga imagen solo para mostrar en pantalla"""
        try:
            imagen = Image.open(ruta)
            imagen = imagen.resize(tamaño, Image.LANCZOS)
            return ImageTk.PhotoImage(imagen)
        except:
            return None


imagen_manager = ImagenManager()


# FUNCIONES DE IMÁGENES
def seleccionar_imagen_libro():
    """Selecciona imagen solo para mostrar"""
    ruta = filedialog.askopenfilename(
        filetypes=[("Imágenes", "*.jpg *.jpeg *.png *.gif")]
    )

    if ruta:
        valido, mensaje = Validaciones.validar_imagen(ruta)
        if valido:
            imagen = imagen_manager.cargar_imagen_para_mostrar(ruta)
            if imagen:
                label_imagen_libro.config(image=imagen)
                label_imagen_libro.image = imagen
                imagen_manager.ruta_actual = ruta
                messagebox.showinfo("Éxito", "Imagen cargada correctamente")
        else:
            messagebox.showerror("Error", mensaje)


def seleccionar_imagen_usuario():
    """Selecciona imagen solo para mostrar"""
    ruta = filedialog.askopenfilename(
        filetypes=[("Imágenes", "*.jpg *.jpeg *.png *.gif")]
    )

    if ruta:
        valido, mensaje = Validaciones.validar_imagen(ruta)
        if valido:
            imagen = imagen_manager.cargar_imagen_para_mostrar(ruta, (120, 120))
            if imagen:
                label_imagen_usuario.config(image=imagen)
                label_imagen_usuario.image = imagen
                imagen_manager.ruta_usuario = ruta
                messagebox.showinfo("Éxito", "Foto cargada correctamente")


def limpiar_imagen_libro():
    """Limpia imagen"""
    label_imagen_libro.config(image='')
    imagen_manager.ruta_actual = None


def limpiar_imagen_usuario():
    """Limpia imagen"""
    label_imagen_usuario.config(image='')
    imagen_manager.ruta_usuario = None


# ========== FUNCIONES PRINCIPALES ==========
def limpiar_libro():
    libro_id.delete(0, tk.END)
    libro_titulo.delete(0, tk.END)
    libro_autor.delete(0, tk.END)
    libro_genero.delete(0, tk.END)
    libro_anio.delete(0, tk.END)
    libro_isbn.delete(0, tk.END)
    limpiar_imagen_libro()


def limpiar_usuario():
    usuario_id.delete(0, tk.END)
    usuario_nombre.delete(0, tk.END)
    usuario_email.delete(0, tk.END)
    usuario_telefono.delete(0, tk.END)
    limpiar_imagen_usuario()


def limpiar_autor():
    autor_id.delete(0, tk.END)
    autor_nombre.delete(0, tk.END)
    autor_nacionalidad.delete(0, tk.END)
    autor_fecha_nacimiento.set_date(datetime.now())


def validar_antes_de_guardar():
    """Validaciones básicas"""
    if not libro_titulo.get().strip():
        messagebox.showerror("Error", "El título es obligatorio")
        return False

    if not libro_autor.get().strip():
        messagebox.showerror("Error", "El autor es obligatorio")
        return False

    año = libro_anio.get().strip()
    if año and not año.isdigit():
        messagebox.showerror("Error", "El año debe ser numérico")
        return False

    return True


def validar_usuario():
    """Validaciones de usuario"""
    if not usuario_nombre.get().strip():
        messagebox.showerror("Error", "El nombre es obligatorio")
        return False

    email = usuario_email.get().strip()
    if not email:
        messagebox.showerror("Error", "El email es obligatorio")
        return False

    if not Validaciones.validar_email(email):
        messagebox.showerror("Error", "Formato de email inválido")
        return False

    telefono = usuario_telefono.get().strip()
    if telefono and not telefono.isdigit():
        messagebox.showerror("Error", "El teléfono debe contener solo números")
        return False

    return True


def guardar_libro():
    if not db.connection and not db.connect():
        return

    if not validar_antes_de_guardar():
        return

    titulo = libro_titulo.get().strip()
    autor = libro_autor.get().strip()

    try:
        if imagen_manager.ruta_actual:
            print("Imagen procesada con PILLOW")

        query = "CALL sp_InsertarLibro(%s, %s, %s, %s, %s)"
        params = (titulo, autor, libro_genero.get().strip(),
                  int(libro_anio.get()) if libro_anio.get().strip() else None,
                  libro_isbn.get().strip())

        success, result = db.execute_query(query, params)

        if success:
            mensaje_sp = result[0][0] if result else "Libro guardado (mensaje no devuelto por SP)"
            messagebox.showinfo("Resultado", mensaje_sp)
            limpiar_libro()
            actualizar_lista_libros()
        else:
            messagebox.showerror("Error", f"Error al guardar: {result}")
    except ValueError:
        messagebox.showerror("Error", "El año debe ser número válido")
    except IndexError:
        messagebox.showerror("Error", "Error inesperado al guardar el libro (índice).")


def guardar_usuario():
    if not db.connection and not db.connect():
        return

    if not validar_usuario():
        return

    nombre = usuario_nombre.get().strip()
    email = usuario_email.get().strip()

    try:
        if hasattr(imagen_manager, 'ruta_usuario') and imagen_manager.ruta_usuario:
            print("Foto procesada con PILLOW")

        query = "CALL sp_InsertarUsuario(%s, %s, %s)"
        params = (nombre, email, usuario_telefono.get().strip() or None)

        success, result = db.execute_query(query, params)

        if success:
            mensaje_sp = result[0][0] if result else "Usuario guardado (mensaje no devuelto por SP)"
            messagebox.showinfo("Resultado", mensaje_sp)
            limpiar_usuario()
            actualizar_lista_usuarios()
        else:
            messagebox.showerror("Error", f"Error al guardar: {result}")
    except Exception as e:
        messagebox.showerror("Error", f"Error inesperado: {e}")


def eliminar_libro():
    if not db.connection and not db.connect():
        return

    id_libro = libro_id.get().strip()

    if not id_libro or not id_libro.isdigit():
        messagebox.showerror("Error", "ID de libro inválido")
        return

    if not messagebox.askyesno("Confirmar", f"¿Eliminar el libro ID {id_libro}?"):
        return

    query = "CALL sp_EliminarLibro(%s)"
    success, result = db.execute_query(query, (int(id_libro),))

    if success:
        mensaje_sp = result[0][0] if result else "Libro eliminado (mensaje no devuelto por SP)"
        messagebox.showinfo("Resultado", mensaje_sp)
        limpiar_libro()
        actualizar_lista_libros()
    else:
        messagebox.showerror("Error", f"Error al eliminar: {result}")


def buscar_libro_por_id():
    if not db.connection and not db.connect():
        return

    id_libro = libro_id.get().strip()

    if not id_libro or not id_libro.isdigit():
        messagebox.showerror("Error", "Ingrese un ID válido")
        return

    query = "SELECT * FROM libros WHERE id = %s"
    success, result = db.execute_query(query, (int(id_libro),))

    if success and result:
        libro = result[0]
        libro_titulo.delete(0, tk.END)
        libro_titulo.insert(0, libro[1])
        libro_autor.delete(0, tk.END)
        libro_autor.insert(0, libro[2])
        libro_genero.delete(0, tk.END)
        libro_genero.insert(0, libro[3] or "")
        libro_anio.delete(0, tk.END)
        if libro[4]:
            libro_anio.insert(0, str(libro[4]))
        libro_isbn.delete(0, tk.END)
        libro_isbn.insert(0, libro[5] or "")
    else:
        messagebox.showinfo("Búsqueda", "Libro no encontrado")


def actualizar_lista_libros():
    if not db.connection and not db.connect():
        return

    for item in tree_libros.get_children():
        tree_libros.delete(item)

    query = "SELECT id, titulo, autor, genero, año_publicacion FROM libros ORDER BY titulo"
    success, result = db.execute_query(query)

    if success:
        for libro in result:
            tree_libros.insert("", tk.END, values=libro)


def actualizar_lista_usuarios():
    if not db.connection and not db.connect():
        return

    for item in tree_usuarios.get_children():
        tree_usuarios.delete(item)

    query = "SELECT id, nombre, email, telefono FROM usuarios ORDER BY nombre"
    success, result = db.execute_query(query)

    if success:
        for usuario in result:
            tree_usuarios.insert("", tk.END, values=usuario)


def realizar_prestamo():
    if not db.connection and not db.connect():
        return

    libro_id_val = prestamo_libro_id.get().strip()
    usuario_id_val = prestamo_usuario_id.get().strip()

    if not libro_id_val or not libro_id_val.isdigit():
        messagebox.showerror("Error", "ID de libro inválido")
        return
    if not usuario_id_val or not usuario_id_val.isdigit():
        messagebox.showerror("Error", "ID de usuario inválido")
        return

    if not messagebox.askyesno("Confirmar", "¿Realizar el préstamo?"):
        return

    query = "CALL sp_RealizarPrestamo(%s, %s)"
    success, result = db.execute_query(query, (int(libro_id_val), int(usuario_id_val)))

    if success:
        mensaje_sp = result[0][0] if result else "Préstamo realizado (mensaje no devuelto por SP)"
        messagebox.showinfo("Resultado", mensaje_sp)
        prestamo_libro_id.delete(0, tk.END)
        prestamo_usuario_id.delete(0, tk.END)
        actualizar_lista_prestamos()
        actualizar_lista_libros()
    else:
        messagebox.showerror("Error", f"Error al realizar préstamo: {result}")


def devolver_libro():
    if not db.connection and not db.connect():
        return

    prestamo_id = devolucion_id.get().strip()

    if not prestamo_id or not prestamo_id.isdigit():
        messagebox.showerror("Error", "ID de préstamo inválido")
        return

    if not messagebox.askyesno("Confirmar", "¿Realizar la devolución?"):
        return

    query = "CALL sp_DevolverLibro(%s)"
    success, result = db.execute_query(query, (int(prestamo_id),))

    if success:
        mensaje_sp = result[0][0] if result else "Libro devuelto (mensaje no devuelto por SP)"
        messagebox.showinfo("Resultado", mensaje_sp)
        devolucion_id.delete(0, tk.END)
        actualizar_lista_prestamos()
        actualizar_lista_libros()
    else:
        messagebox.showerror("Error", f"Error al devolver libro: {result}")


def actualizar_lista_prestamos():
    if not db.connection and not db.connect():
        return

    for item in tree_prestamos.get_children():
        tree_prestamos.delete(item)

    query = """SELECT p.id, l.titulo, u.nombre, p.fecha_prestamo, 
               CASE WHEN p.devuelto THEN 'Sí' ELSE 'No' END
               FROM prestamos p
               JOIN libros l ON p.libro_id = l.id
               JOIN usuarios u ON p.usuario_id = u.id
               ORDER BY p.fecha_prestamo DESC"""
    success, result = db.execute_query(query)

    if success:
        for prestamo in result:
            tree_prestamos.insert("", tk.END, values=prestamo)


def guardar_autor():
    if not db.connection and not db.connect():
        return

    nombre = autor_nombre.get().strip()

    if not nombre:
        messagebox.showerror("Error", "El nombre del autor es obligatorio")
        return

    query = "CALL sp_InsertarAutor(%s, %s, %s)"
    params = (nombre, autor_nacionalidad.get().strip() or None,
              autor_fecha_nacimiento.get_date())

    success, result = db.execute_query(query, params)

    if success:
        mensaje_sp = result[0][0] if result else "Autor guardado (mensaje no devuelto por SP)"
        messagebox.showinfo("Resultado", mensaje_sp)
        limpiar_autor()
        actualizar_lista_autores()
    else:
        messagebox.showerror("Error", f"Error al guardar: {result}")


def eliminar_autor():
    if not db.connection and not db.connect():
        return

    id_autor = autor_id.get().strip()

    if not id_autor or not id_autor.isdigit():
        messagebox.showerror("Error", "ID de autor inválido")
        return

    if not messagebox.askyesno("Confirmar", f"¿Eliminar el autor ID {id_autor}?"):
        return

    query = "CALL sp_EliminarAutor(%s)"
    success, result = db.execute_query(query, (int(id_autor),))

    if success:
        mensaje_sp = result[0][0] if result else "Autor eliminado"
        messagebox.showinfo("Resultado", mensaje_sp)
        limpiar_autor()
        actualizar_lista_autores()
    else:
        messagebox.showerror("Error", f"Error al eliminar: {result}")


def actualizar_lista_autores():
    if not db.connection and not db.connect():
        return

    for item in tree_autores.get_children():
        tree_autores.delete(item)

    query = "SELECT id, nombre, nacionalidad, fecha_nacimiento FROM autores ORDER BY nombre"
    success, result = db.execute_query(query)

    if success:
        for autor in result:
            fecha_formateada = autor[3].strftime("%d/%m/%Y") if autor[3] else ""
            tree_autores.insert("", tk.END, values=(autor[0], autor[1], autor[2] or "", fecha_formateada))


#  EXPORTACIONES - VERSIÓN CORREGIDA
def exportar_a_excel(datos, nombre_archivo, encabezados):
    """Exporta datos a un archivo Excel usando openpyxl"""
    if not OPENPYXL_DISPONIBLE:
        messagebox.showerror("Error", "Instala 'openpyxl': pip install openpyxl")
        return

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = nombre_archivo.replace('.xlsx', '')

        # Agregar encabezados
        for col_num, encabezado in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_num, value=encabezado)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Agregar datos
        for row_num, fila in enumerate(datos, start=2):
            for col_num, valor in enumerate(fila, start=1):
                ws.cell(row=row_num, column=col_num, value=str(valor))

        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(nombre_archivo)
        messagebox.showinfo("Éxito", f"Datos exportados a {nombre_archivo}")
    except Exception as e:
        messagebox.showerror("Error", f"Error al exportar a Excel: {e}")


def exportar_a_pdf(datos, nombre_archivo, encabezados, titulo_tabla="Datos"):
    """Exporta datos a un archivo PDF usando reportlab"""
    if not REPORTLAB_DISPONIBLE:
        messagebox.showerror("Error", "Instala 'reportlab': pip install reportlab")
        return

    try:
        doc = SimpleDocTemplate(nombre_archivo, pagesize=A4)
        elements = []

        # Estilo de título
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1,
        )
        title = Paragraph(titulo_tabla, title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))

        # Crear tabla
        data_for_table = [encabezados] + datos
        table = Table(data_for_table)

        # Estilo de la tabla
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        elements.append(table)
        doc.build(elements)
        messagebox.showinfo("Éxito", f"Datos exportados a {nombre_archivo}")
    except Exception as e:
        messagebox.showerror("Error", f"Error al exportar a PDF: {e}")


def exportar_libros_excel():
    if not db.connection and not db.connect():
        return

    query = "SELECT id, titulo, autor, genero, año_publicacion, isbn, disponible FROM libros ORDER BY titulo"
    success, result = db.execute_query(query)

    if success:
        encabezados = ["ID", "Título", "Autor", "Género", "Año", "ISBN", "Disponible"]
        exportar_a_excel(result, "libros_exportados.xlsx", encabezados)
    else:
        messagebox.showerror("Error", "No se pudieron obtener los datos de libros")


def exportar_libros_pdf():
    if not db.connection and not db.connect():
        return

    query = "SELECT id, titulo, autor, genero, año_publicacion, isbn, disponible FROM libros ORDER BY titulo"
    success, result = db.execute_query(query)

    if success:
        encabezados = ["ID", "Título", "Autor", "Género", "Año", "ISBN", "Disponible"]
        exportar_a_pdf(result, "libros_exportados.pdf", encabezados, "Listado de Libros")
    else:
        messagebox.showerror("Error", "No se pudieron obtener los datos de libros")


def exportar_usuarios_excel():
    if not db.connection and not db.connect():
        return

    query = "SELECT id, nombre, email, telefono FROM usuarios ORDER BY nombre"
    success, result = db.execute_query(query)

    if success:
        encabezados = ["ID", "Nombre", "Email", "Teléfono"]
        exportar_a_excel(result, "usuarios_exportados.xlsx", encabezados)
    else:
        messagebox.showerror("Error", "No se pudieron obtener los datos de usuarios")


def exportar_usuarios_pdf():
    if not db.connection and not db.connect():
        return

    query = "SELECT id, nombre, email, telefono FROM usuarios ORDER BY nombre"
    success, result = db.execute_query(query)

    if success:
        encabezados = ["ID", "Nombre", "Email", "Teléfono"]
        exportar_a_pdf(result, "usuarios_exportados.pdf", encabezados, "Listado de Usuarios")
    else:
        messagebox.showerror("Error", "No se pudieron obtener los datos de usuarios")


def exportar_prestamos_excel():
    if not db.connection and not db.connect():
        return

    query = """SELECT p.id, l.titulo, u.nombre, p.fecha_prestamo, p.fecha_devolucion, 
               CASE WHEN p.devuelto THEN 'Sí' ELSE 'No' END
               FROM prestamos p
               JOIN libros l ON p.libro_id = l.id
               JOIN usuarios u ON p.usuario_id = u.id
               ORDER BY p.fecha_prestamo DESC"""
    success, result = db.execute_query(query)

    if success:
        encabezados = ["ID Préstamo", "Libro", "Usuario", "Fecha Préstamo", "Fecha Devolución", "Devuelto"]
        exportar_a_excel(result, "prestamos_exportados.xlsx", encabezados)
    else:
        messagebox.showerror("Error", "No se pudieron obtener los datos de préstamos")


def exportar_prestamos_pdf():
    if not db.connection and not db.connect():
        return

    query = """SELECT p.id, l.titulo, u.nombre, p.fecha_prestamo, p.fecha_devolucion, 
               CASE WHEN p.devuelto THEN 'Sí' ELSE 'No' END
               FROM prestamos p
               JOIN libros l ON p.libro_id = l.id
               JOIN usuarios u ON p.usuario_id = u.id
               ORDER BY p.fecha_prestamo DESC"""
    success, result = db.execute_query(query)

    if success:
        encabezados = ["ID Préstamo", "Libro", "Usuario", "Fecha Préstamo", "Fecha Devolución", "Devuelto"]
        exportar_a_pdf(result, "prestamos_exportados.pdf", encabezados, "Historial de Préstamos")
    else:
        messagebox.showerror("Error", "No se pudieron obtener los datos de préstamos")


def exportar_autores_excel():
    if not db.connection and not db.connect():
        return

    query = "SELECT id, nombre, nacionalidad, fecha_nacimiento FROM autores ORDER BY nombre"
    success, result = db.execute_query(query)

    if success:
        encabezados = ["ID", "Nombre", "Nacionalidad", "Fecha Nacimiento"]
        exportar_a_excel(result, "autores_exportados.xlsx", encabezados)
    else:
        messagebox.showerror("Error", "No se pudieron obtener los datos de autores")


def exportar_autores_pdf():
    if not db.connection and not db.connect():
        return

    query = "SELECT id, nombre, nacionalidad, fecha_nacimiento FROM autores ORDER BY nombre"
    success, result = db.execute_query(query)

    if success:
        encabezados = ["ID", "Nombre", "Nacionalidad", "Fecha Nacimiento"]
        exportar_a_pdf(result, "autores_exportados.pdf", encabezados, "Listado de Autores")
    else:
        messagebox.showerror("Error", "No se pudieron obtener los datos de autores")


# INTERFAZ GRÁFICA (el resto del código permanece igual)
root = tk.Tk()
root.title("Sistema de Gestión de Biblioteca")
root.geometry("1200x800")
root.configure(bg='white')

try:
    root.iconbitmap("favicon.ico")
except:
    pass

# Crear pestañas
notebook = ttk.Notebook(root)

# Pestaña Libros
tab_libros = ttk.Frame(notebook)
notebook.add(tab_libros, text="Libros")

# Pestaña Usuarios
tab_usuarios = ttk.Frame(notebook)
notebook.add(tab_usuarios, text="Usuarios")

# Pestaña Préstamos
tab_prestamos = ttk.Frame(notebook)
notebook.add(tab_prestamos, text="Préstamos")

# Pestaña Autores
tab_autores = ttk.Frame(notebook)
notebook.add(tab_autores, text="Autores")

notebook.pack(expand=True, fill="both", padx=10, pady=5)

# INTERFAZ LIBROS
frame_form_libro = ttk.LabelFrame(tab_libros, text="Gestión de Libros", padding=10)
frame_form_libro.pack(fill="x", padx=10, pady=5)

# Formulario
frame_izq = ttk.Frame(frame_form_libro)
frame_izq.pack(side="left", fill="both", expand=True)

ttk.Label(frame_izq, text="ID:").grid(row=0, column=0, padx=5, pady=2, sticky="w")
libro_id = ttk.Entry(frame_izq, width=10)
libro_id.grid(row=0, column=1, padx=5, pady=2)
ttk.Button(frame_izq, text="Buscar por ID", command=buscar_libro_por_id).grid(row=0, column=2, padx=5)

ttk.Label(frame_izq, text="Título:*").grid(row=1, column=0, padx=5, pady=2, sticky="w")
libro_titulo = ttk.Entry(frame_izq, width=30)
libro_titulo.grid(row=1, column=1, padx=5, pady=2, columnspan=2)

ttk.Label(frame_izq, text="Autor:*").grid(row=2, column=0, padx=5, pady=2, sticky="w")
libro_autor = ttk.Entry(frame_izq, width=30)
libro_autor.grid(row=2, column=1, padx=5, pady=2, columnspan=2)

ttk.Label(frame_izq, text="Género:").grid(row=3, column=0, padx=5, pady=2, sticky="w")
libro_genero = ttk.Entry(frame_izq, width=30)
libro_genero.grid(row=3, column=1, padx=5, pady=2, columnspan=2)

ttk.Label(frame_izq, text="Año:").grid(row=4, column=0, padx=5, pady=2, sticky="w")
libro_anio = ttk.Entry(frame_izq, width=10)
libro_anio.grid(row=4, column=1, padx=5, pady=2, sticky="w")
vcmd = (root.register(Validaciones.solo_numeros), '%P')
libro_anio.configure(validate="key", validatecommand=vcmd)

ttk.Label(frame_izq, text="ISBN:").grid(row=4, column=2, padx=5, pady=2, sticky="w")
libro_isbn = ttk.Entry(frame_izq, width=20)
libro_isbn.grid(row=4, column=3, padx=5, pady=2)

# Imagen
frame_der = ttk.Frame(frame_form_libro)
frame_der.pack(side="right", padx=20)

ttk.Label(frame_der, text="Portada del Libro").pack()
label_imagen_libro = tk.Label(frame_der, background="lightgray", width=20, height=10)
label_imagen_libro.pack(pady=5)
frame_botones_img = ttk.Frame(frame_der)
frame_botones_img.pack()
ttk.Button(frame_botones_img, text="Seleccionar Imagen",
           command=seleccionar_imagen_libro).pack(side="left", padx=2)
ttk.Button(frame_botones_img, text="Limpiar Imagen",
           command=limpiar_imagen_libro).pack(side="left", padx=2)

# Botones
frame_botones = ttk.Frame(frame_izq)
frame_botones.grid(row=5, column=0, columnspan=4, pady=10)
ttk.Button(frame_botones, text="Guardar", command=guardar_libro).pack(side="left", padx=5)
ttk.Button(frame_botones, text="Eliminar", command=eliminar_libro).pack(side="left", padx=5)
ttk.Button(frame_botones, text="Limpiar", command=limpiar_libro).pack(side="left", padx=5)
# Botones de exportación para Libros
ttk.Button(frame_botones, text="Exportar a Excel", command=exportar_libros_excel).pack(side="left", padx=5)
ttk.Button(frame_botones, text="Exportar a PDF", command=exportar_libros_pdf).pack(side="left", padx=5)

# Lista
frame_lista_libros = ttk.LabelFrame(tab_libros, text="Lista de Libros", padding=10)
frame_lista_libros.pack(fill="both", expand=True, padx=10, pady=5)
columns_libros = ("ID", "Título", "Autor", "Género", "Año")
tree_libros = ttk.Treeview(frame_lista_libros, columns=columns_libros, show="headings", height=12)
for col in columns_libros:
    tree_libros.heading(col, text=col)
tree_libros.column("ID", width=50)
tree_libros.column("Título", width=250)
tree_libros.column("Autor", width=150)
tree_libros.column("Género", width=100)
tree_libros.column("Año", width=80)
tree_libros.pack(fill="both", expand=True)

# INTERFAZ USUARIOS
frame_form_usuario = ttk.LabelFrame(tab_usuarios, text="Gestión de Usuarios", padding=10)
frame_form_usuario.pack(fill="x", padx=10, p